import pymysql
import openpyxl
import os

path=os.path.dirname(os.path.abspath(__file__))

#数据源渲染配置项
def setting_detail(datatype):
    if datatype=='mysql':
        data=[
            {'title':'主机','name':'mysql_host','type':'input'},
            {'title':'端口','name':'mysql_port','type':'input'},
            {'title':'用户名','name':'mysql_usr','type':'input'},
            {'title':'密码','name':'mysql_pwd','type':'pwd'},
            {'title':'库','name':'mysql_db','type':'input'},
            {'title':'表','name':'mysql_tab','type':'input'}
        ]

    elif datatype=='xlsx':
        files=os.listdir(path+'/xlsx_files')
        fdatas=[x for x in files if '.xlsx' in x]
        data=[
            {
                'title':'choose_file',
                'name':'choose_file',
                'type':'select_rt',
                'func':'show_sheets',
                'datas':['']+fdatas
                }
        ]
    return data
#mysql加载tables(未应用)
def mysql_show_tbs(dbInfo):
    db=pymysql.connect(host=dbInfo['host'],port=dbInfo['port'],user=dbInfo['usr'],password=dbInfo['pwd'],database=dbInfo['db']) #本地登录
    cur=db.cursor()
    sql='''
        show tables
    '''
    cur.execute(sql)
    result=cur.fetchall()
    db.close()
    return [x[0] for x in result]
#mysql加载字段
def mysql_show_cols(dbInfo):
    db=pymysql.connect(host=dbInfo['host'],port=dbInfo['port'],user=dbInfo['usr'],password=dbInfo['pwd'],database=dbInfo['db']) #本地登录
    cur=db.cursor()
    sql='''
        desc {tabname}
    '''.format(tabname=dbInfo['tb'])
    cur.execute(sql)
    result=cur.fetchall()
    db.close()
    return [{'name':x[0],'value':x[0]} for x in result]
#创建sql
def create_sql(require_data):
    c_ts=["sum","mid","avg","max","min","count"]
    r_ts=[">",">=","<","<=","=","<>","in"]
    n_s=require_data['names']
    c_s=[c_ts[int(x['action'])]+'('+x['name']+')' for x in require_data['caculations']]
    r_s=[]
    for mulit_rules in require_data['rules']:
        multi_rs=[]
        for rule_raw in mulit_rules:
            action=r_ts[int(rule_raw['action'])]
            if action=='in':
                query=rule_raw['name']+' in ('+rule_raw['action_value']+')'
            elif action in ['=','<>']:
                query=rule_raw['name']+action+'\''+rule_raw['action_value']+'\''
            else:
                query=rule_raw['name']+action+rule_raw['action_value']
            multi_rs.append(query)
        multi_query='('+' and '.join(multi_rs)+')'
        r_s.append(multi_query)


    if len(n_s)>0:
        name_query=','.join(n_s)
        gb_query=' group by '+','.join(n_s)
    else:
        name_query=''
        gb_query=''
    if len(c_s)>0:
        cal_query=','.join(c_s)
    else:
        cal_query=''
    if len(r_s)>0:
        rule_query='where '+' or '.join(r_s)
    else:
        rule_query=''
    if name_query!='' and cal_query!='':
        select_q=name_query+','+cal_query
    else:
        select_q=name_query+cal_query

    sql='select '+select_q+' from {tabname} '+rule_query+gb_query+' limit 1000'
    return sql
#mysql取数
def mysql_search(dbInfo,sql):
    sql_c=sql.format(tabname=dbInfo['tb'])
    db=pymysql.connect(host=dbInfo['host'],port=dbInfo['port'],user=dbInfo['usr'],password=dbInfo['pwd'],database=dbInfo['db']) #本地登录
    cur=db.cursor()
    cur.execute(sql_c)
    result=cur.fetchall()
    db.close()
    return [[str(y) for y in x] for x in result]
#xlsx获取文件
def sheetname(fname):
    f=openpyxl.load_workbook(path+'/xlsx_files/'+fname,read_only=True)
    datas=f.sheetnames
    data=[
            {
                'title':'choose_sheet',
                'name':'choose_sheet',
                'type':'select',
                'datas':datas
                }
        ]
    return data
#xlsx获取加载工作表字段
def xlsx_columns(fname,sheetname):
    f=openpyxl.load_workbook(path+'/xlsx_files/'+fname,read_only=True)
    ws=f[sheetname]
    maxc=ws.max_column
    return [{'name':ws.cell(1,x+1).value,'value':x} for x in range(maxc)]

#xlsx处理
def xlsx_rule(rules,dfr):
    raw_multi_fliters=[]
    for multi_rule in rules:
        fliter_rows=[]
        if len(multi_rule)>0:
            for rule_dic in multi_rule:
                cid=int(rule_dic['name'])
                mthd=rule_dic['action']
                if mthd=='6':
                    m_value=rule_dic['action_value'].split(',')
                elif mthd in['0','1','2','3']:
                    m_value=float(rule_dic['action_value'])
                else:
                    m_value=rule_dic['action_value']
                for dr in range(len(dfr)):
                    cell=dfr[dr][cid]
                    if mthd=='0':
                        if cell<=m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='1':
                        if cell<m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='2':
                        if cell>=m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='3':
                        if cell>m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='4':
                        if cell!=m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='5':
                        if cell==m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
                    elif mthd=='6':
                        if cell not in m_value and dr not in fliter_rows:
                            fliter_rows.append(dr)
        raw_multi_fliters.append([x for x in range(len(dfr)) if x not in fliter_rows])
    after_rule_rows=[]
    for rmf in raw_multi_fliters:
        for rf in rmf:
            if rf not in after_rule_rows:
                after_rule_rows.append(rf)
    return after_rule_rows

def xlsx_produce(fn,sn,rqsdata):
    f=openpyxl.load_workbook(path+'/xlsx_files/'+fn,read_only=True)
    ws=f[sn]
    names=rqsdata['names']
    caculations=rqsdata['caculations']
    if len(names)>0:
        namecolids=[int(x) for x in names]
    else:
        namecolids=[]
    if len(caculations)>0:
        calcolids=[int(x['name']) for x in caculations]
        actionids=[int(x['action']) for x in caculations]
    else:
        calcolids=[]
    colidx=namecolids+calcolids
    rules=rqsdata['rules']
    maxc=ws.max_column
    maxr=ws.max_row
    #读取sheet后关闭减少读写文件
    dfr=[]
    cols_raw=[ws.cell(1,x+1).value for x in range(maxc)]
    cols=[]
    cal_types=["sum","mid","avg","max","min","count"]
    for nameid in namecolids:
        cols.append(cols_raw[nameid])
    for calcolid in range(len(calcolids)):
        cols.append(str(cols_raw[calcolid])+'_'+cal_types[actionids[calcolid]])
    data_matrix=[cols]
    
    for r in range(maxr-1):
        rdata=[ws.cell(r+2,x+1).value for x in range(maxc)]
        if rdata[0]!=None:
            dfr.append(rdata)
    f.close()

    #筛选缩小df
    ar_rows=xlsx_rule(rules,dfr)
    df=[[dfr[x][y] for y in colidx] for x in range(len(dfr)) if x in ar_rows]

    #合并names列中相同字段
    if len(namecolids)>0:
        rows=[]
        for nid in range(len(namecolids)):
            rs=[]
            ndatas=[]
            for r in range(len(df)):
                value=df[r][nid]
                if value not in ndatas:
                    ndatas.append(value)
                    rs.append(r)
            for ri in rs:
                if ri not in rows:
                    rows.append(ri)
    else:
        rows=[x for x in range(len(df))]
    
    #去重字段矩阵
    result_df=[]
    if len(namecolids)>0:
        for row in rows:
            include_rows=[]
            colnames=df[row][:len(namecolids)]
            #namedatas.append(colnames)
            for dfrow in range(len(df)):
                if df[dfrow][:len(namecolids)]==colnames:
                    include_rows.append(dfrow)
            calrows=[]
            if len(calcolids)>0:
                for calid in range(len(calcolids)):
                    calcoldata_s=[]
                    calcol=caculations[calid]
                    for in_row in include_rows:
                        calcoldata_s.append(df[in_row][len(namecolids)+calid])
                    calcoldatas=sorted(calcoldata_s,reverse=True)
                    if calcol['action']=='0':
                        caldata=sum(calcoldatas)
                    elif calcol['action']=='1':
                        collen=len(calcoldatas)
                        if collen%2==0:
                            caldata=(calcoldatas[collen//2]+calcoldatas[collen//2+1])/2
                        else:
                            caldata=calcoldatas[collen//2+1]
                    elif calcol['action']=='2' and len(calcoldatas)>0:
                        caldata=sum(calcoldatas)/len(calcoldatas)
                    elif calcol['action']=='3':
                        caldata=max(calcoldatas)
                    elif calcol['action']=='4':
                        caldata=min(calcoldatas)
                    elif calcol['action']=='5':
                        caldata=len(calcoldatas)
                    else:
                        caldata=0
                    calrows.append(caldata)
            result_df.append(colnames+calrows)
    else:
        calrows=[]
        if len(calcolids)>0:
            for calid in range(len(calcolids)):
                calcoldata_s=[]
                calcol=caculations[calid]
                for in_row in rows:
                    calcoldata_s.append(df[in_row][len(namecolids)+calid])
                calcoldatas=sorted(calcoldata_s,reverse=True)
                if calcol['action']=='0':
                    print(calcoldatas)
                    caldata=sum(calcoldatas)
                elif calcol['action']=='1':
                    collen=len(calcoldatas)
                    if collen%2==0:
                        caldata=(calcoldatas[collen//2]+calcoldatas[collen//2+1])/2
                    else:
                        caldata=calcoldatas[collen//2+1]
                elif calcol['action']=='2' and len(calcoldatas)>0:
                    caldata=sum(calcoldatas)/len(calcoldatas)
                elif calcol['action']=='3':
                    caldata=max(calcoldatas)
                elif calcol['action']=='4':
                    caldata=min(calcoldatas)
                elif calcol['action']=='5':
                    caldata=len(calcoldatas)
                else:
                    caldata=0
                calrows.append(caldata)
        result_df.append(calrows)
    data_matrix+=result_df
    return data_matrix

